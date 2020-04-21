# -*- coding: utf-8 -*-

import openpyxl
import numpy as np

path = r'C:\Users\sammy\Documents\GitHub\Data_Tools\AWQMS_Parsers'

#%% Get Station IDs in Region
wb = openpyxl.load_workbook(filename = path + '\DEQ_Stations.xlsx')

sheet_ranges = wb['DEQ_Stations_No_Duplicates']

monitoring_IDs = []
for i in range(3,sheet_ranges.max_row):
    if sheet_ranges['B'+str(i)].value is not None:
        monitoring_IDs.append(sheet_ranges['B'+str(i)].value)
        
wb.close()   
 
#%%   
wb = openpyxl.load_workbook(filename = path + '\DEQ_Stations_Standard_Report_9505.xlsx')

ws = wb['Results']

rows = ws.iter_rows(min_row=2, max_row=2) # returns a generator of rows
first_row = next(rows) # get the first row

#Columns with Station Info
col_station_IDs = [c.column for c in first_row if c.value == 'Monitoring Location ID'][0] 
col_station_coords = [c.column for c in first_row if c.value == 'Monitoring Location Latitude'][0]

#Columns of Samples
date_col = [c.column for c in first_row if c.value == 'Activity Start Date'][0]
sample_name_col = [c.column for c in first_row if c.value == 'Characteristic Name'][0]
sample_value_col = [c.column for c in first_row if c.value == 'Result Value'][0]
water_type_col = [c.column for c in first_row if c.value == 'Activity Media Subdivision'][0]

#%% Read Standard Report

station_data = dict()

for i in range(3,ws.max_row):
    std_id = ws.cell(row=i, column=col_station_IDs).value
    if std_id in monitoring_IDs:
        if str(std_id) not in station_data.keys():
            station_data[str(std_id)] = dict()
            station_data[str(std_id)]['Data'] = dict()
            station_data[str(std_id)]['Coords'] = [float(ws.cell(row=i, column=col_station_coords).value),float(ws.cell(row=i, column=col_station_coords+1).value)]
            station_data[str(std_id)]['water_type'] = str(ws.cell(row=i, column=water_type_col).value)
        
        var_name = str(ws.cell(row=i, column=sample_name_col).value)
        year_str = ws.cell(row=i, column=date_col).value.year
        mon_str = ws.cell(row=i, column=date_col).value.month
        day_str = ws.cell(row=i, column=date_col).value.day
        
        if var_name not in station_data[str(std_id)]['Data'].keys():
            station_data[str(std_id)]['Data'][var_name] = dict()
            station_data[str(std_id)]['Data'][var_name]['Time'] = []
            station_data[str(std_id)]['Data'][var_name]['Value'] = []
            station_data[str(std_id)]['Data'][var_name]['Units'] = []
            
        station_data[str(std_id)]['Data'][var_name]['Time'].append([year_str,mon_str,day_str])
        
        if '>' in ws.cell(row=i, column=sample_value_col).value:
            station_data[str(std_id)]['Data'][var_name]['Value'].append(float(ws.cell(row=i, column=sample_value_col).value.strip('>'))+1)
        elif '<' in ws.cell(row=i, column=sample_value_col).value:
            station_data[str(std_id)]['Data'][var_name]['Value'].append(float(ws.cell(row=i, column=sample_value_col).value.strip('<'))-1)
        elif 'ND' in ws.cell(row=i, column=sample_value_col).value:
            station_data[str(std_id)]['Data'][var_name]['Value'].append(np.float('NaN'))
        else:
            station_data[str(std_id)]['Data'][var_name]['Value'].append(float(ws.cell(row=i, column=sample_value_col).value))
            
        station_data[str(std_id)]['Data'][var_name]['Units'].append(str(ws.cell(row=i, column=sample_value_col+1).value))
                     
wb.close()  

#%% Missing Stations

missing_stations = []
for i in monitoring_IDs:
    if i not in station_data.keys():
        missing_stations.append(i)

#%% Data Stats

water_type_num_rec = dict()
for i in station_data.keys():
    if station_data[i]['water_type'] not in water_type_num_rec.keys():
        water_type_num_rec[station_data[i]['water_type']] = dict()
        water_type_num_rec[station_data[i]['water_type']]['Num_Records'] = 0
        water_type_num_rec[station_data[i]['water_type']]['Variables'] = dict()
        
    water_type_num_rec[station_data[i]['water_type']]['Num_Records'] += 1
    
    for j in station_data[i]['Data'].keys():
        if j not in water_type_num_rec[station_data[i]['water_type']]['Variables'].keys():
            water_type_num_rec[station_data[i]['water_type']]['Variables'][j] = dict()
            water_type_num_rec[station_data[i]['water_type']]['Variables'][j]['Total_Stations'] = 0
            water_type_num_rec[station_data[i]['water_type']]['Variables'][j]['Avg_Obs'] = []

        water_type_num_rec[station_data[i]['water_type']]['Variables'][j]['Total_Stations'] += 1
        
        temp_counter = 0
        for jj in station_data[i]['Data'][j]['Value']:
            temp_counter += 1
        
        water_type_num_rec[station_data[i]['water_type']]['Variables'][j]['Avg_Obs'].append(temp_counter)
          
